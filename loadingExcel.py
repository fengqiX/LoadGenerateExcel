import os,sys
from openpyxl import load_workbook


def load(path,fileName):
    #checking directory & file
    try:
        os.chdir(path)
    except:
        print("path is invalid")
    finally:
        print(os.getcwd())
    try:
        wb = load_workbook(fileName,data_only=True)
    except:
        print("file does not exist!")
    #open the first sheet     
    sheets = wb.sheetnames;
    sheet=wb.get_sheet_by_name(sheets[0])
    
    dataset = {}
    name=""
    tempValue=list()
    for col in sheet.columns:
        tempValue=list()
        for cell in col:
            if cell.row==1:
                name=cell.value
            else:
                tempValue.append(cell.value)
        dataset[name]=tempValue
   # print(dataset)
    return dataset

if __name__ == '__main__':
    load(path,fileName)
