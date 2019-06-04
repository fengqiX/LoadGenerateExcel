import os,sys,math
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side

#change directory
def generate_unlock(dict):
    matching = {
        "MRBTS":["MRBTS ID",],
        "LTBTS":["LNBTS ID",],
        "LCELL":["LNCEL ID",],
        "name":["LNCEL Name",],
        "earfcnDL":["DL EARFCN [#]",],
        "earfcnUL":["UL EARFCN [#]",]
    }
    path=""
    fileNameTemp="Cells to Unlock Unreserve.xlsx"
    count=0
    sitecode="123 "
    try:
        wb = load_workbook(fileNameTemp,data_only=True)
    except:
        print("file does not exist!")
    #adding style
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Ebrima',size=8)
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)  
    #get index based on the field name
    for key,value in matching.items():
        if value[0] in dict["Names"]:
            matching.get(key).append(dict["Names"].index(value[0]))
    #get data based on index from dict
    for key,value in matching.items():
        temp=list()
        for data in dict["Data"]:
            temp.append(data[value[1]])
        value.append(temp)
        count=len(temp)
    #loading data to spreadsheet
    ws=wb["LNCEL"]    
    for col in ws.iter_cols(min_col=2,max_col=7,min_row=2,max_row=count+2):
        key=""
        i=0
        for cell in col:
            if cell.row==2:
                key=cell.value
                i=0
                print(key)
            else:
                try:
                    cell.value=matching[key][2][i]
                    cell.style=highlight
                except:
                    print(matching[key])
                i+=1
    #final step-save file and return
    wb.save(sitecode+fileNameTemp)
    return
        
"""
latitudeSign	orientationOfMajorAxis	uncertaintyAltitude	uncertaintySemiMajor	

"""
def generate_CM(dict):
    matching = {
        "LNCEL Name":["LNCEL Name",],
        "Sector Id Letter":["Sector Id Letter",],
        "Sector Id Number":["Sector Id Number"],
        "LNBTS Name":[{"concatenate":{"name":"Property Name","content":"L"}}],
        "Property Name":["Property Name"],
        "LTE SectorCount within Site":["LTE SectorCount within Site"],
        "LTE SectorCount within eNodeB":["LTE SectorCount within eNodeB"],
        "MRBTS ID":["MRBTS ID",],
        "LNBTS ID":["LNBTS ID",],
        "LNCEL ID":["LNCEL ID",],
        "Latitude":["Latitude [Dec Deg]"],
        "Longitude":["Longitude [Dec Deg]"],
        "antennaOrientation":[{"default":0}],
        "altitude":[{"default":6}],
        "confidence":[{"default":0}],
        "degreesOfLatitude":[{"calculation":"Latitude"}], #=ROUND((2^23)/90*Latitude,0)
        #round(math.pow(2,23)/90*,0)
        "degreesOfLongitude":[{"calculation":"Longitude"}], #=ROUND(((2^24)/360*M4),0)
        "directionOfAltitude":[{"default":1}],
        "latitudeSign":[{"default":1}],
        "orientationOfMajorAxis":[{"default":0}],
        "uncertaintyAltitude":[{"default":0}],
        "uncertaintySemiMajor":[{"default":0}],
        "uncertaintySemiMinor":[{"default":0}]
    }
    path=""
    fileNameTemp="CM Name Upload.xlsx"
    count=0
    sitecode="123 "
    try:
        wb = load_workbook(fileNameTemp,data_only=True)
    except:
        print("file does not exist!")
        
    #adding style
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Ebrima',size=8)
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)  
    
    #get index based on the field name
    for key,value in matching.items():
        if len(value)>0 and type(value[0])==str:
            if value[0] in dict["Names"]:
                matching.get(key).append(dict["Names"].index(value[0]))
    #get data based on index from dict
    for key,value in matching.items():
        if len(value)>0 and type(value[0])==str:
            temp=list()
            for data in dict["Data"]:
                temp.append(data[value[1]])
            value.append(temp)
            count=len(temp)
    #loading data to spreadsheet
    ws=wb["CM Name"]    
    #load indentical data from original Excel file 
    for col in ws.iter_cols(min_col=2,max_col=24,min_row=2,max_row=count+2):
        key=""
        i=0
        tempValue=""
        for cell in col:
            if cell.row==2:
                key=cell.value
                i=0
              #  print(key)
            else:
                try:
                    if type(matching[key][0])==str:
                       # print("str works")
                        cell.value=matching[key][2][i]
                        cell.style=highlight
                except:
                    print(key)
                    print("Unexpected error:", sys.exc_info()[0])
                i+=1
    #load data which needs some calculation            
    for col in ws.iter_cols(min_col=2,max_col=24,min_row=2,max_row=count+2):
        key=""
        i=0
        tempValue=""
        for cell in col:
            if cell.row==2:
                key=cell.value
                i=0
            else:
                if type(matching[key][0]).__name__=='dict':
                    #print("dict works")
                    if "default" in matching[key][0]:
                        cell.value=matching[key][0]["default"]
                        cell.style=highlight
                    if "concatenate" in matching[key][0]:
                        tempValue=matching[matching[key][0]["concatenate"]["name"]][2][i]+matching[key][0]["concatenate"]["content"]
                        cell.value=tempValue
                        cell.style=highlight
                    if "calculation" in matching[key][0]:
                        tempValue=float(matching[matching[key][0]["calculation"]][2][i])
                        if matching[key][0]["calculation"]=="Latitude":
                            cell.value=round(math.pow(2,23)/90*tempValue,0)
                        if matching[key][0]["calculation"]=="Longitude":
                            cell.value=round(math.pow(2,24)/360*tempValue,0)
                        cell.style=highlight
                    i+=1

    #final step-save file and return
    wb.save(sitecode+fileNameTemp)
    return
























