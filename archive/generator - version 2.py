import os,sys,math
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side,Alignment

#change directory
def generate_unlock(dict):
    matching = {
        "MRBTS":{"CIQ_name":"MRBTS ID","index":int(),"data":[]},
        "LTBTS":{"CIQ_name":"LNBTS ID","index":int(),"data":[]},
        "LCELL":{"CIQ_name":"LNCEL ID","index":int(),"data":[]},
        "name":{"CIQ_name":"LNCEL Name","index":int(),"data":[]},
        "earfcnDL":{"CIQ_name":"DL EARFCN [#]","index":int(),"data":[]},
        "earfcnUL":{"CIQ_name":"UL EARFCN [#]","index":int(),"data":[]}
    }
    path=""
    fileNameTemp="Cells to Unlock Unreserve.xlsx"
    count=0
    sitecode="123 "
    try:
        wb = load_workbook(fileNameTemp, data_only=True)
    except:
        print("file does not exist!")
    #adding style
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Ebrima',size=8,)
    highlight.alignment=Alignment(horizontal='center')
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)  
    #get index based on the field name
    for key,value in matching.items():
        if value["CIQ_name"] in dict["Names"]:
            #matching.get(key).append(dict["Names"].index(value[0]))
            matching[key]["index"]=dict["Names"].index(value["CIQ_name"])
    #get data based on index from dict
    for key,value in matching.items():
        temp=list()
        for data in dict["Data"]:
            value["data"].append(data[value["index"]])
        count=len(value["data"])
    #loading data to spreadsheet
    print(len(matching))
    ws=wb["LNCEL"]    
    for col in ws.iter_cols(min_col=2,max_col=len(matching)+1,min_row=2,max_row=count+2):
        key=""
        i=0
        for cell in col:
            if cell.row==2:
                key=cell.value
                i=0
                print(key)
            else:
                cell.value=matching[key]["data"][i]
                cell.style=highlight
                i+=1
    #final step-save file and return
    wb.save(sitecode+fileNameTemp)
    print("Unlock & Unreserve file is generated!")
    return
        
"""
"""
def generate_CM(dict):
    matching = {
        "LNCEL Name":                   {"CIQ_name":"LNCEL Name","index":int(),"data":[]},
        "Sector Id Letter":             {"CIQ_name":"Sector Id Letter","index":int(),"data":[]},
        "Sector Id Number":             {"CIQ_name":"Sector Id Number","index":int(),"data":[]},
        "LNBTS Name":                   {"CIQ_name":"none",
                                         "alt_name":"concatenate",
                                         "concatenate":{"name":"Property Name",
                                                        "content":"L"}
                                        },
        "Property Name":                {"CIQ_name":"Property Name","index":int(),"data":[]},
        "LTE SectorCount within Site":  {"CIQ_name":"LTE SectorCount within Site","index":int(),"data":[]},
        "LTE SectorCount within eNodeB":{"CIQ_name":"LTE SectorCount within eNodeB","index":int(),"data":[]},
        "MRBTS ID":                     {"CIQ_name":"MRBTS ID","index":int(),"data":[]},
        "LNBTS ID":                     {"CIQ_name":"LNBTS ID","index":int(),"data":[]},
        "LNCEL ID":                     {"CIQ_name":"LNCEL ID","index":int(),"data":[]},
        "Latitude":                     {"CIQ_name":"Latitude [Dec Deg]","index":int(),"data":[]},
        "Longitude":                    {"CIQ_name":"Longitude [Dec Deg]","index":int(),"data":[]},
        "antennaOrientation":           {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":0},
        "altitude":                     {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":6}, #also found 7 in ACAa028
        "confidence":                   {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":0},
        "degreesOfLatitude":            {"CIQ_name":"none",
                                         "alt_name":"calculation",
                                         "calculation":"Latitude"}, #=ROUND((2^23)/90*Latitude,0)
        "degreesOfLongitude":           {"CIQ_name":"none",
                                         "alt_name":"calculation",
                                         "calculation":"Longitude"}, #=ROUND(((2^24)/360*M4),0)
        "directionOfAltitude":          {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":1},
        "latitudeSign":                 {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":1},
        "orientationOfMajorAxis":       {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":0},
        "uncertaintyAltitude":          {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":0},
        "uncertaintySemiMajor":         {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":0},
        "uncertaintySemiMinor":         {"CIQ_name":"none",
                                         "alt_name":"default",
                                         "default":0}
    }
    path=""
    fileNameTemp="CM Name Upload.xlsx"
    count=0
    sitecode="123 "
    try:
        wb = load_workbook(fileNameTemp,data_only=True)
    except:
        print("file does not exist!")
        
    #adding style
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Ebrima',size=8)
    highlight.alignment=Alignment(horizontal='center')
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)  
    
   #get index based on the field name
    for key,value in matching.items():
        if value["CIQ_name"] in dict["Names"]:
            #matching.get(key).append(dict["Names"].index(value[0]))
            matching[key]["index"]=dict["Names"].index(value["CIQ_name"])
    #get data based on index from dict
    #print(dict)
    #print(matching)
    for key,value in matching.items():
        temp=list()
        for data in dict["Data"]:
            if "data" in value:
                value["data"].append(data[value["index"]])
    count = len(matching["LNCEL Name"]["data"])
    #loading data to spreadsheet
    ws=wb["CM Name"]    
    #load indentical data from original Excel file 
    for col in ws.iter_cols(min_col=2,max_col=len(matching)+1,min_row=2,max_row=count+2):
        key=""
        i=0
        tempValue=""
        for cell in col:
            if cell.row==2:
                key=cell.value
                i=0
              #  print(key)
            else:
                if matching[key]["CIQ_name"]!="none":
                   # print("str works")
                    cell.value=matching[key]["data"][i]
                    cell.style=highlight
                i+=1
    #load data which needs some calculation            
    for col in ws.iter_cols(min_col=2,max_col=len(matching)+1,min_row=2,max_row=count+2):
        key=""
        i=0
        tempValue=""
        for cell in col:
            if cell.row==2:
                key=cell.value
                i=0
            else:
                if matching[key]["CIQ_name"]=="none":
                    #print("dict works")
                    if matching[key]["alt_name"] == "default":
                        cell.value=matching[key]["default"]
                        cell.style=highlight
                    if matching[key]["alt_name"] == "concatenate":
                        tempValue=matching[matching[key]["concatenate"]["name"]]["data"][i]+matching[key]["concatenate"]["content"]
                        cell.value=tempValue
                        cell.style=highlight
                    if matching[key]["alt_name"] == "calculation":
                        tempValue=float(matching[matching[key]["calculation"]]["data"][i])
                        if matching[key]["calculation"]=="Latitude":
                            cell.value=round(math.pow(2,23)/90*tempValue,0)
                        if matching[key]["calculation"]=="Longitude":
                            cell.value=round(math.pow(2,24)/360*tempValue,0)
                        cell.style=highlight
                    i+=1

    #final step-save file and return
    wb.save(sitecode+fileNameTemp)
    print("CM file is generated!")
    return

"""                                
""" 
def generate_SAS_info(dict):
    matching_E911={"Name":          {"CIQ_name":"Property Name","index":int(),"data":[]},
                   "Sector Name":   {"CIQ_name":"none",
                                     "concatenate":"Sector",
                                     "data":[]},
                   "Cell Name":     {"CIQ_name":"LNCEL Name","index":int(),"data":[]},
                   "LNBTS ID":      {"CIQ_name":"LNBTS ID","index":int(),"data":[]},
                   "LNCELl ID":     {"CIQ_name":"LNCEL ID","index":int(),"data":[]},
                   "Cell ID":       {"CIQ_name":"none",
                                     "concatenate":"Cell",
                                     "data":[]},
                   "ECGI":          {"CIQ_name":"none",
                                     "calculation":"ECGI",
                                     "data":[]},
                   "PCI":           {"CIQ_name":"PCI","index":int(),"data":[]},
                   "Latitude":      {"CIQ_name":"Latitude [Dec Deg]","index":int(),"data":[]},
                   "Longitude":     {"CIQ_name":"Longitude [Dec Deg]","index":int(),"data":[]},
                   "cell_azimuth":  {"CIQ_name":"none","default":0},
                   "Street_Name":   {"CIQ_name":"none"},
                   "Municipality":  {"CIQ_name":"none"},
                   "Province":      {"CIQ_name":"none"},
                   "Site_Market":   {"CIQ_name":"none"},
                   "Postal_Code":   {"CIQ_name":"none"},
                   "Tested":        {"CIQ_name":"none","default":0},
                   "Ready":         {"CIQ_name":"none","default":0},
                   "cell_technology":{"CIQ_name":"none"},
                   "band":          {"CIQ_name":"Band Indicator","index":int(),"data":[]},
                   "TAC":           {"CIQ_name":"TAC","index":int(),"data":[]},
                   "Rac":           {"CIQ_name":"none","default":1},
                   "Rnc":           {"CIQ_name":"none","default":31}	
                  }
                                 
    matching_VAS={"MCC":                        {"CIQ_name":"MCC","index":int(),"data":[]},
                  "MNC":                        {"CIQ_name":"MNC","index":int(),"data":[]}, 
                  "Cell-ID (ECGI)":             {"CIQ_name":"none","alter":matching_E911["ECGI"]},
                  "CellName":                   {"CIQ_name":"LNCEL Name","index":int(),"data":[]},
                  "CellSite":                   {"CIQ_name":"Property Name","index":int(),"data":[]},
                  "PhysicalCellID (ECGI)":      {"CIQ_name":"none","alter":matching_E911["PCI"]},
                  "Latitude":                   {"CIQ_name":"Latitude [Dec Deg]","index":int(),"data":[]},
                  "Longitude":                  {"CIQ_name":"Longitude [Dec Deg]","index":int(),"data":[]},
                  "GroundHeight":               {"CIQ_name":"none","default":722},
                  "Orientation":                {"CIQ_name":"none","default":0},
                  "Opening":                    {"CIQ_name":"none","default":360},
                  "Range":                      {"CIQ_name":"none","default":100},
                  "ServingAltitudeUncertainty": {"CIQ_name":"none","default":12},
                  "SupportedPDEs":              {"CIQ_name":"none","default":2},   
                  "NeighbourList":              {"CIQ_name":"none"},
                  "SIPRoute":                   {"CIQ_name":"none"},
                  "MiscFlags":                  {"CIQ_name":"none","default":1},
                  "eNodeBToAntennaDelay":       {"CIQ_name":"none"},
                  "note":                       {"CIQ_name":"none"},
                  "date":                       {"CIQ_name":"none"}
                 }
    matching_Core={"eNodeB Name":               {"CIQ_name":"none",      
                                                "concatenate":"eNodeB",
                                                 "data":[]},
                   "TAC":                       {"CIQ_name":"TAC","index":int(),"data":[]},
                   "E-UTRAN cell identity (ECI)":{"CIQ_name":"none","alter":matching_E911["ECGI"]},
                   "Mapping cell name":         {"CIQ_name":"LNCEL Name","index":int(),"data":[]},
                   "Mapping cell SAC":          {"CIQ_name":"LNCEL ID","index":int(),"data":[]},
                   "Mapping cell LAC":          {"CIQ_name":"none","calculation":"Mapping cell LAC","data":[]}, #not working
                   "Address":                   {"CIQ_name":"none"},
                   "City":                      {"CIQ_name":"none"}
                    } 
    path=""
    fileNameTemp="LTE E911 SAS Info.xlsx"
    count=0
    sitecode="123 "
    try:
        wb = load_workbook(fileNameTemp,data_only=True)
    except:
        print("file does not exist!")
        
    #adding style
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Ebrima',size=8)
    highlight.alignment=Alignment(horizontal='center')
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)  
    

    #working on the first worksheet                               
    ws=wb["E911-LTE"]   
    #get index based on the field name
    for key,value in matching_E911.items():
        if value["CIQ_name"] in dict["Names"]:
            matching_E911[key]["index"]=dict["Names"].index(value["CIQ_name"])
    for key,value in matching_E911.items():
        temp=list()
        for data in dict["Data"]:
            try:
                if "index" in value:
                    value["data"].append(data[value["index"]])
            except:
                print(key)
    count = len(matching_E911["Name"]["data"])
    #load indentical data from original Excel file 
    for col in ws.iter_cols(min_col=1,max_col=len(matching_E911),min_row=1,max_row=count+1):
        key=""
        i=0
        tempValue=""
        for cell in col:
            if cell.row==1:
                key=cell.value
                i=0
              #  print(key)
            else:
                if matching_E911[key]["CIQ_name"]!="none":
                   # print("str works")
                    cell.value=matching_E911[key]["data"][i]
                    cell.style=highlight
                i+=1     
  #load data which needs some calculation            
    for col in ws.iter_cols(min_col=1,max_col=len(matching_E911),min_row=1,max_row=count+1):
        key=""
        i=0
        tempValue=""
        for cell in col:
            if cell.row==1:
                key=cell.value
                i=0
            else:
                if matching_E911[key]["CIQ_name"]=="none":
                    #print("dict works")
                    tempValue
                    if "default" in matching_E911[key]:
                        tempValue=matching_E911[key]["default"]
                    if "concatenate" in matching_E911[key]:
                        if matching_E911[key]["concatenate"]=="Sector":
                            tempValue= str(matching_E911["Name"]["data"][i])+"-"+str(matching_E911["LNBTS ID"]["data"][i])+"_S1"
                            matching_E911[key]["data"].append(tempValue)
                        if matching_E911[key]["concatenate"]=="Cell":
                            tempValue= str(matching_E911["LNBTS ID"]["data"][i])+"_"+str(matching_E911["LNCELl ID"]["data"][i])
                            matching_E911[key]["data"].append(tempValue)                       
                    if "calculation" in matching_E911[key]:
                        #tempValue=int()
                        if  matching_E911[key]["calculation"]=="ECGI":
                            tempValue=int(matching_E911["LNBTS ID"]["data"][i])*256+int(matching_E911["LNCELl ID"]["data"][i])
                            matching_E911[key]["data"].append(tempValue) 
                    cell.value=tempValue
                    cell.style=highlight    
                    i+=1
#working on the second worksheet                               
    #working on the second worksheet 
    ws=wb["VAS-LTE"]    
    #get index based on the field name
    for key,value in matching_VAS.items():
        if value["CIQ_name"] in dict["Names"]:
            matching_VAS[key]["index"]=dict["Names"].index(value["CIQ_name"])
    for key,value in matching_VAS.items():
        temp=list()
        for data in dict["Data"]:
            if "index" in value:
                value["data"].append(data[value["index"]])
    #load indentical data from original Excel file 
    for col in ws.iter_cols(min_col=2,max_col=len(matching_VAS)+1,min_row=1,max_row=count+1):
        key=""
        i=0
        tempValue=""
        for cell in col:
            if cell.row==1:
                key=cell.value
                i=0
              #  print(key)
            else:
                if matching_VAS[key]["CIQ_name"]!="none":
                   # print("str works")
                    cell.value=matching_VAS[key]["data"][i]
                    cell.style=highlight
                i+=1     
  #load data which needs some calculation            
    for col in ws.iter_cols(min_col=2,max_col=len(matching_VAS)+1,min_row=1,max_row=count+1):
        key=""
        i=0
        for cell in col:
            if cell.row==1:
                key=cell.value
                i=0
            else:
                if matching_VAS[key]["CIQ_name"]=="none":
                    #print("dict works")
                    tempValue=""
                    if "alter" in matching_VAS[key]:
                        tempValue=matching_VAS[key]["alter"]["data"][i]
                    if "default" in matching_VAS[key]:
                        tempValue=matching_VAS[key]["default"]
                    cell.value=tempValue
                    cell.style=highlight    
                    i+=1
    #working on the third worksheet 
    ws=wb["Core-LTE"]    
    #get index based on the field name
    for key,value in matching_Core.items():
        if value["CIQ_name"] in dict["Names"]:
            matching_Core[key]["index"]=dict["Names"].index(value["CIQ_name"])
    for key,value in matching_Core.items():
        temp=list()
        for data in dict["Data"]:
            try:    
                if "index" in value:
                    value["data"].append(data[value["index"]])
            except:
                print(key)
    #load indentical data from original Excel file 
    for col in ws.iter_cols(min_col=1,max_col=len(matching_Core),min_row=1,max_row=count+1):
        key=""
        i=0
        for cell in col:
            if cell.row==1:
                key=cell.value
                i=0
              #  print(key)
            else:
                if matching_Core[key]["CIQ_name"]!="none":
                   # print("str works")
                    cell.value=matching_Core[key]["data"][i]
                    cell.style=highlight
                i+=1     
  #load data which needs some calculation            
    for col in ws.iter_cols(min_col=1,max_col=len(matching_Core),min_row=1,max_row=count+1):
        key=""
        i=0
        for cell in col:
            if cell.row==1:
                key=cell.value
                i=0
            else:
                if matching_Core[key]["CIQ_name"]=="none":
                    #print("dict works")
                    tempValue=""
                    if "alter" in matching_Core[key]:
                        tempValue=matching_Core[key]["alter"]["data"][i]
                    if "concatenate" in matching_Core[key]:
                        if matching_Core[key]["concatenate"]=="eNodeB":
                            tempValue= str(matching_E911["Name"]["data"][i])+"-"+str(matching_E911["LNBTS ID"]["data"][i])
                            matching_Core[key]["data"].append(tempValue)               
                    if "calculation" in matching_Core[key]:
                        #tempValue=int()
                        if  matching_Core[key]["calculation"]=="Mapping cell LAC":
                            tempValue="2"+str(matching_Core["TAC"]["data"][i])[1:]
                            
                    cell.value=tempValue
                    cell.style=highlight    
                    i+=1
    #final step-save file and return
    wb.save(sitecode+fileNameTemp)
    print("SAS file is generated!")
    return


if __name__ == '__main__':
    generate_CM(dict)
    generate_unlock(dict)
    generate_SAS_info(dict)