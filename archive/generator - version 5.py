import os,sys,math,time
import re,json
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
# address={
#     "fullAddress":"",
#     "partialAddress":"",
#     "city":"",
#     "province":"",
#     "abbreviation":"",
#     "zipcode":""
# }
__addresslist = []
matching_UN = {"MRBTS":{"CIQ_name":"MRBTS ID","data":[]},
    "LTBTS":{"CIQ_name":"LNBTS ID","data":[]},
    "LCELL":{"CIQ_name":"LNCEL ID","data":[]},
    "name":{"CIQ_name":"LNCEL Name","data":[]},
    "earfcnDL":{"CIQ_name":"DL EARFCN [#]","data":[]},
    "earfcnUL":{"CIQ_name":"UL EARFCN [#]","data":[]}}
matching_CM = {
        "LNCEL Name":                   {"CIQ_name":"LNCEL Name","data":[]},
        "Sector Id Letter":             {"CIQ_name":"Sector Id Letter","data":[]},
        "Sector Id Number":             {"CIQ_name":"Sector Id Number","data":[]},
        "LNBTS Name":                   {"CIQ_name":"LNBTS Name","data":[]},
        "Property Name":                {"CIQ_name":"Property Name","data":[]},
        "LTE SectorCount within Site":  {"CIQ_name":"LTE SectorCount within Site","data":[]},
        "LTE SectorCount within eNodeB":{"CIQ_name":"LTE SectorCount within eNodeB","data":[]},
        "MRBTS ID":                     {"CIQ_name":"MRBTS ID","data":[]},
        "LNBTS ID":                     {"CIQ_name":"LNBTS ID","data":[]},
        "LNCEL ID":                     {"CIQ_name":"LNCEL ID","data":[]},
        "Latitude":                     {"CIQ_name":"Latitude [Dec Deg]", "data":[]},
        "Longitude":                    {"CIQ_name":"Longitude [Dec Deg]", "data":[]},
        "antennaOrientation":           {"CIQ_name":"none","default":0},
        "altitude":                     {"CIQ_name":"none","default":6}, #also found 7 in ACAa028
        "confidence":                   {"CIQ_name":"none","default":0},
        "degreesOfLatitude":            {"CIQ_name":"none","calculation":"CM Latitude"}, #=ROUND((2^23)/90*Latitude,0)
        "degreesOfLongitude":           {"CIQ_name":"none","calculation":"CM Longitude"}, #=ROUND(((2^24)/360*Longitude),0)
        "directionOfAltitude":          {"CIQ_name":"none","default":1},
        "latitudeSign":                 {"CIQ_name":"none","default":1},
        "orientationOfMajorAxis":       {"CIQ_name":"none","default":0},
        "uncertaintyAltitude":          {"CIQ_name":"none","default":0},
        "uncertaintySemiMajor":         {"CIQ_name":"none","default":0},
        "uncertaintySemiMinor":         {"CIQ_name":"none","default":0}
    }
matching_E911={"Name":          {"CIQ_name":"Property Name", "data":[]},
               "Sector Name":   {"CIQ_name":"none",
                                 "concatenate":"E911 Sector Name",
                                 "data":[]},
               "Cell Name":     {"CIQ_name":"LNCEL Name", "data":[]},
               "LNBTS ID":      {"CIQ_name":"LNBTS ID", "data":[]},
               "LNCELl ID":     {"CIQ_name":"LNCEL ID", "data":[]},
               "Cell ID":       {"CIQ_name":"none",
                                 "concatenate":"E911 Cell ID",
                                 "data":[]},
               "ECGI":          {"CIQ_name":"none",
                                 "calculation":"E911 ECGI",
                                 "data":[]},
               "PCI":           {"CIQ_name":"PCI", "data":[]},
               "Latitude":      {"CIQ_name":"Latitude [Dec Deg]", "data":[]},
               "Longitude":     {"CIQ_name":"Longitude [Dec Deg]", "data":[]},
               "cell_azimuth":  {"CIQ_name":"none","default":0},
               "Street_Name":   {"CIQ_name":"none","address":"partialAddress"},
               "Municipality":  {"CIQ_name":"none"},
               "Province":      {"CIQ_name":"none","address":"province"},
               "Site_Market":   {"CIQ_name":"none","address":"city"},
               "Postal_Code":   {"CIQ_name":"none","address":"zipcode"},
               "Tested":        {"CIQ_name":"none","default":0},
               "Ready":         {"CIQ_name":"none","default":0},
               "cell_technology":{"CIQ_name":"none",
                                 "pairing":{"B66":"LTE-F1","B4":"LTE-F2","B7":"LTE-F3","B46":"LTE-F4"}},
               "band":          {"CIQ_name":"Band Indicator", "data":[]},
               "TAC":           {"CIQ_name":"TAC", "data":[]},
               "Rac":           {"CIQ_name":"none","default":1},
               "Rnc":           {"CIQ_name":"none","default":31}	
              }                     
matching_VAS={"MCC":                        {"CIQ_name":"MCC", "data":[]},
              "MNC":                        {"CIQ_name":"MNC", "data":[]}, 
              "Cell-ID (ECGI)":             {"CIQ_name":"none","alteration":matching_E911["ECGI"]},
              "CellName":                   {"CIQ_name":"LNCEL Name","data":[]},
              "CellSite":                   {"CIQ_name":"Property Name","data":[]},
              "PhysicalCellID (ECGI)":      {"CIQ_name":"PCI","data":[]},
              "Latitude":                   {"CIQ_name":"Latitude [Dec Deg]", "data":[]},
              "Longitude":                  {"CIQ_name":"Longitude [Dec Deg]", "data":[]},
              "GroundHeight":               {"CIQ_name":"none"},
              "Orientation":                {"CIQ_name":"none","default":0},
              "Opening":                    {"CIQ_name":"none","default":360},
              "Range":                      {"CIQ_name":"none","default":100},
              "ServingAltitudeUncertainty": {"CIQ_name":"none","default":12},
              "SupportedPDEs":              {"CIQ_name":"none","default":2},   
              "NeighbourList":              {"CIQ_name":"none"},
              "SIPRoute":                   {"CIQ_name":"none"},
              "MiscFlags":                  {"CIQ_name":"none","default":1},
              "eNodeBToAntennaDelay":       {"CIQ_name":"none"},
              "note":                       {"CIQ_name":"none"},
              "date":                       {"CIQ_name":"none"}
             }
matching_Core={"eNodeB Name":               {"CIQ_name":"none",      
                                                "concatenate":"Core eNodeB",
                                                 "data":[]},
                   "TAC":                       {"CIQ_name":"TAC", "data":[]},
                   "E-UTRAN cell identity (ECI)":{"CIQ_name":"none","alteration":matching_E911["ECGI"]},
                   "Mapping cell name":         {"CIQ_name":"LNCEL Name", "data":[]},
                   "Mapping cell SAC":          {"CIQ_name":"LNCEL ID", "data":[]},
                   "Mapping cell LAC":          {"CIQ_name":"none","calculation":"Core Mapping cell LAC","data":[]}, 
                   "Address":                   {"CIQ_name":"Property Address","data":[]},
                   "City":                      {"CIQ_name":"none","address":"city"}
                    } 


def __addressAnalyzer(address_str):
    try:
        with open('zip2prov.json', 'r') as f:
            data_zip = json.load(f)
    except Exception as result:
        print('Unexpected Error {}'.format(result))

    pattern_ZIP = r"[A-Z]\d[A-Z] \d[A-Z]\d"
    pattern_Province_A=r"\b[A-Z]{2}\b"
    
    str = "2390 47 Ave Calgary,AB  T2T 5W5"
    #address_str=str
    address["fullAddress"]=address_str

    match_zip=re.search(pattern_ZIP,address_str)
    if match_zip:
        address["zipcode"]=re.search(pattern_ZIP,address_str).group()
    else: print("no matching! ZIP CODE")

    match_Province_A = re.search(pattern_Province_A,address_str)
    if match_Province_A:
        address["abbreviation"]=match_Province_A.group()
    else: 
        print("no matching! PROVINCE")
    if address["zipcode"] != "":
        for province in data_zip:
            if address["zipcode"][0] in province["ZFL"]:
                if province["Abbreviation"].upper() != address["abbreviation"].upper():
                    address["abbreviation"]=province["Abbreviation"].upper()
                address["province"]=province["Province"]
                for city in province["Cities"]:
                    if city.lower() in address_str.lower():
                        address["city"]=city
                        address["partialAddress"]=address_str[:address_str.find(city)]           
    return
#    print(address)
    
def __cell_style():
    #adding style
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Ebrima',size=8,)
    highlight.alignment=Alignment(horizontal='center')
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return highlight

def __loading_data(dataset, matching):
    """
    """
    for key,value in matching.items():
        if value["CIQ_name"] in dataset:
            matching[key]["data"]=dataset.get(value["CIQ_name"]).copy()
    return matching

def __uploading_data(ws,minCol,maxCol,minRow,maxRow,matching,highlight):

    for col in ws.iter_cols(min_col=minCol,max_col=maxCol,min_row=minRow,max_row=maxRow):
        key=""
        i=0
        for cell in col:
            if cell.row==minRow:
                key=cell.value
                i=0
                #print(key)
            else:
                if matching[key]["CIQ_name"]!="none":   
                    cell.value=matching[key]["data"][i]
                    cell.style=highlight
                i+=1
    for col in ws.iter_cols(min_col=minCol,max_col=maxCol,min_row=minRow,max_row=maxRow):
        key=""
        i=0
        for cell in col:
            if cell.row==minRow:
                key=cell.value
                i=0
            else:
                tempValue=""
                if matching[key]["CIQ_name"]=="none":
                    #print("dict works")
                    if "default" in matching[key]:
                       #print(matching[key]["default"])
                        tempValue=matching[key]["default"]
                    elif "concatenate" in matching[key]:
                        # if matching[key]["concatenate"]=="CM LNBTS Name":
                        #     tempValue=str(matching["Property Name"]["data"][i])+"L"
                        if matching[key]["concatenate"]=="E911 Sector Name":
                            tempValue= str(matching["Name"]["data"][i])+"-"+str(matching["LNBTS ID"]["data"][i])+"_S1"
                        if matching[key]["concatenate"]=="E911 Cell ID":
                            tempValue= str(matching["LNBTS ID"]["data"][i])+"_"+str(matching["LNCELl ID"]["data"][i]) 
                        if matching[key]["concatenate"]=="Core eNodeB":
                            tempValue= str(matching_E911["Name"]["data"][i])+"-"+str(matching_E911["LNBTS ID"]["data"][i])
                    elif "calculation" in matching[key]:
                        if matching[key]["calculation"]=="CM Latitude":
                            tempValue=float(matching["Latitude"]["data"][i])
                            tempValue=round(math.pow(2,23)/90*tempValue,0)               
                        if matching[key]["calculation"]=="CM Longitude":
                            tempValue=float(matching["Longitude"]["data"][i])
                            tempValue=round(math.pow(2,24)/360*tempValue,0)
                        if  matching[key]["calculation"]=="E911 ECGI":
                            tempValue=int(matching["LNBTS ID"]["data"][i])*256+int(matching["LNCELl ID"]["data"][i])
                            matching[key]["data"].append(tempValue)
                        if  matching[key]["calculation"]=="Core Mapping cell LAC":
                            tempValue=int("2"+str(matching["TAC"]["data"][i])[1:])
                    elif "alteration" in matching[key]:
                        #print(matching[key]["alteration"])
                        tempValue=matching[key]["alteration"]["data"][i]
                    elif "pairing" in matching[key]:
                        tempValue=matching[key]["pairing"][matching["band"]["data"][i]]
                    elif "address" in matching[key]:
                        # address_str = matching["Street_Name"]["data"][i]
                        # __addressAnalyzer(address_str)
                        tempValue=__addresslist[i][matching[key]["address"]]
                    cell.value=tempValue
                    cell.style=highlight
                    i+=1

        

#change directory
def __editingDate():
    return time.strftime("%d-%b-%Y", time.localtime())

def generate_unlock(dataset,sitecode=""):
    path=""
    fileNameTemp="Cells to Unlock Unreserve.xlsx"
    count=0
    #sitecode="AEDa015 "
    try:
        wb = load_workbook(fileNameTemp, data_only=True)
    except:
        print("file does not exist!")
    
    highlight=__cell_style()
    wb.add_named_style(highlight)     
    __loading_data(dataset, matching_UN)
    #print(matching)
    ws=wb["LNCEL"]    
    count=len(matching_UN["MRBTS"]["data"])
    __uploading_data(ws,2,len(matching_UN)+1,2,count+2,matching_UN,highlight)
    #final step-save file and return
    #print(matching)
    wb.save(sitecode+fileNameTemp)
    print("Unlock & Unreserve file is generated!")
    return
        
"""
"""
def generate_CM(dataset,sitecode=""):
    
    path=""
    fileNameTemp="CM Name Upload.xlsx"
    count=0
    #sitecode="AEDa015 "
    try:
        wb = load_workbook(fileNameTemp,data_only=True)
    except:
        print("file does not exist!")
    
    highlight=__cell_style()
    wb.add_named_style(highlight)     
    __loading_data(dataset, matching_CM)
    
   #get index based on the field name

    count = len(matching_CM["LNCEL Name"]["data"])
    ws=wb["CM Name"]  
    __uploading_data(ws,2,len(matching_CM)+1,2,count+2,matching_CM,highlight)
    #final step-save file and return
    wb.save(sitecode+fileNameTemp)
    print("CM file is generated!")
    return

"""                                
""" 
def generate_SAS_info(dataset,sitecode=""):
  
    path=""
    fileNameTemp="LTE E911 SAS Info.xlsx"
    count=0
    #sitecode="AEDa015 "
    try:
        wb = load_workbook(fileNameTemp,data_only=True)
    except:
        print("file does not exist!")
        
    #adding style
    highlight=__cell_style()
    wb.add_named_style(highlight)     
    
    #get address details
    # address_str=dataset["Property Address"][0]
    # __addressAnalyzer(address_str)
    #print(address)
    __addresslist = dataset['Analyzed Address']
    #load data from CIQ to local storage
    __loading_data(dataset, matching_E911)
    __loading_data(dataset, matching_VAS)
    __loading_data(dataset, matching_Core)
    
    #working on the first worksheet                               
    ws=wb["E911-LTE"]   
    count = len(matching_E911["Name"]["data"])
    __uploading_data(ws,1,len(matching_E911),1,count+1,matching_E911,highlight)
                                  
    #working on the second worksheet 
    ws=wb["VAS-LTE"]
    __uploading_data(ws,2,len(matching_VAS)+1,1,count+1,matching_VAS,highlight)
 
    #working on the third worksheet 
    ws=wb["Core-LTE"]      
    __uploading_data(ws,1,len(matching_Core),1,count+1,matching_Core,highlight)

    #final step-save file and return
    wb.save(sitecode+fileNameTemp)
    print("SAS file is generated!")
    return


def generate_SCF(dataset,sitecode=""):
    matching_SCF={
        "Site ID":{"CIQ_name":"MRBTS ID", "data":[]},
        "WBTS ID":{"CIQ_name":"LNBTS ID", "data":[]},
        "WBTS name":{"CIQ_name":"Property Name", "data":[]},
    }
    path=""
    fileNameTemp="Small Cell SCF Flex.xlsx"
    count=0
    #sitecode="AEDa015 "
    try:
        wb = load_workbook(fileNameTemp,data_only=True)
    except:
        print("file does not exist!")
    i=0
    count=0
    for x in dataset[matching_SCF["Site ID"]["CIQ_name"]]:
        if x not in matching_SCF["Site ID"]["data"]:
            matching_SCF["Site ID"]["data"].append(x)
            matching_SCF["WBTS ID"]["data"].append(dataset[matching_SCF["WBTS ID"]["CIQ_name"]][i])
            matching_SCF["WBTS name"]["data"].append(dataset[matching_SCF["WBTS name"]["CIQ_name"]][i])
            count+=1
        i+=1
    
    print(count)
    print(matching_SCF)
    ws=wb["Note"]
    ws['A15']=__editingDate()
    for col in ws.iter_cols(min_col=2,max_col=4,min_row=14,max_row=count+14):
        key=""
        j=0
        for cell in col:
            if cell.row==14:
                key=cell.value
            else:
                cell.value=matching_SCF[key]["data"][j]
                j+=1
    ws=wb["IPNB"]
    col_IPNBid = 2
    col_IpbasedRouteName=19
    for x in range(9,9+count):
        ws.cell(row=x, column=col_IPNBid, value=matching_SCF["Site ID"]["data"][x-9])
        ws.cell(row=x, column=col_IpbasedRouteName, value=matching_SCF["WBTS name"]["data"][x-9])
    wb.save(sitecode+fileNameTemp)
    print("SCF file is generated!")
    
if __name__ == '__main__':
    generate_CM(dataset)
    generate_unlock(dataset)
    generate_SAS_info(dataset)
    generate_SCF(dataset)