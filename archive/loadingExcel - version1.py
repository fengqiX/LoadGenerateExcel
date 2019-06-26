import os,sys
from openpyxl import load_workbook


def load(path,fileName):
    try:
        os.chdir(path)
    except:
        print("path is invalid")
    finally:
        print(os.getcwd())
    try:
        wb = load_workbook(fileName,data_only=True)
    except:
        print("file does not exist!")
        
    sheets = wb.sheetnames;
  #  print(sheets)
    sheet=wb.get_sheet_by_name(sheets[0])
  #  print(sheet)
 #   print(list(sheet.rows)[0])
    namelist = list()
    for cell in list(sheet.rows)[0]:
        namelist.append(cell.value)
    content = list()
    temp = list()
    for row in list(sheet.rows)[1:]:
        temp = list()
        for cell in row:
            temp.append(cell.value or "")
        content.append(temp)
    dict={
        'Names':namelist,
        'Data':content
    }
    return dict
  #  for n in names:
 #       print(n.value)
#    print(len(list(sheet.rows)))
#   for cell in list(sheet.rows)[2]:
#       print(cell.value)